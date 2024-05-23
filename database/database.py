#database.py
import psycopg2
from psycopg2.extras import RealDictCursor
from config.config import config
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from fpdf import FPDF
import matplotlib.pyplot as plt
import numpy as np


# Функция для получения соединения с базой данных
def get_db_connection():
    connection = psycopg2.connect(
        dbname=config.DB_NAME,
        user=config.DB_USER,
        password=config.DB_PASSWORD,
        host=config.DB_HOST,
        port=config.DB_PORT
    )
    return connection


# Функция для проверки закрытия семестра
def is_semester_closed():
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(
        "SELECT semester_closed FROM student_semester_courses LIMIT 1")  # Запрос статуса семестра
    result = cursor.fetchone()  # Получение результата
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения
    return result['semester_closed'] if result else False   # Возврат статуса семестра или False


# Функция для получения оценок студентов
def fetch_grades(assessment_number):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(f"""
        SELECT s.student_id, s.surname, s.name, g.course_id, g.assessment{assessment_number}, c.code, c.name as course_name
        FROM students s
        JOIN grades g ON s.student_id = g.student_id
        JOIN courses c ON g.course_id = c.course_id
    """)    # Запрос оценок студентов
    grades = cursor.fetchall()  # Получение всех результатов
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения
    return grades   # Возврат списка оценок


# Функция для генерации отчета по аттестации
def generate_assessment_report(assessment_number):
    grades = fetch_grades(assessment_number)    # Получение оценок студентов
    workbook = openpyxl.Workbook()  # Создание нового рабочей книги
    sheet = workbook.active # Получение активного листа
    sheet.title = f'Assessment {assessment_number} Report'  # Установка названия листа

    # Заголовки
    headers = ['№', 'ФИО', 'Курс']
    courses = list(set([f"{grade['code']} – {grade['course_name']}" for grade in grades]))  # Получение уникальных курсов
    headers.extend(courses)
    sheet.append(headers)   # Добавление заголовков на лист

    # Стиль заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))

    for cell in sheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    # Заполнение данных студентов
    student_data = {}
    for idx, grade in enumerate(grades):
        student_id = grade['student_id']
        course_name = f"{grade['code']} – {grade['course_name']}"
        if student_id not in student_data:
            student_data[student_id] = {
                'ФИО': f"{grade['surname']} {grade['name']}",
                'Курс': '4',
                'Оценки': {course: '' for course in courses}
            }
        student_data[student_id]['Оценки'][course_name] = grade[f'assessment{assessment_number}']

    row_num = 2
    for student in student_data.values():
        row = [row_num - 1, student['ФИО'], student['Курс']]
        row.extend([student['Оценки'][course] for course in courses])
        sheet.append(row)

        # Стиль данных
        for col_num in range(1, len(row) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

            if col_num == 2 and any(isinstance(grade, int) and grade <= 15 for grade in student['Оценки'].values()):
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF")

            if col_num > 3 and isinstance(cell.value, int) and cell.value <= 15:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF")

        # Установка цвета фона для строк
        if row_num % 2 == 0:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for col_num in range(1, len(row) + 1):
            if not (col_num == 2 and any(
                    isinstance(grade, int) and grade <= 15 for grade in student['Оценки'].values())) and not (
                    col_num > 3 and isinstance(sheet.cell(row=row_num, column=col_num).value, int) and sheet.cell(
                    row=row_num, column=col_num).value <= 15):
                sheet.cell(row=row_num, column=col_num).fill = fill

        row_num += 1

    # Авторазмер колонок
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Генерация графиков по каждому курсу
    for idx, course in enumerate(courses):
        chart_data = {
            'ВСЕГО': 0,
            'МЕНЬШЕ 15* (недопуск)': 0,
            'ОТ 15 ДО 25': 0,
            'БОЛЬШЕ 25': 0
        }
        for student in student_data.values():
            score = student['Оценки'][course]
            if isinstance(score, int):
                chart_data['ВСЕГО'] += 1
                if score <= 15:
                    chart_data['МЕНЬШЕ 15* (недопуск)'] += 1
                elif score <= 25:
                    chart_data['ОТ 15 ДО 25'] += 1
                else:
                    chart_data['БОЛЬШЕ 25'] += 1

        data = [
            ['ВСЕГО', chart_data['ВСЕГО']],
            ['МЕНЬШЕ 15* (недопуск)', chart_data['МЕНЬШЕ 15* (недопуск)']],
            ['ОТ 15 ДО 25', chart_data['ОТ 15 ДО 25']],
            ['БОЛЬШЕ 25', chart_data['БОЛЬШЕ 25']]
        ]

        chart_sheet = workbook.create_sheet(title=f"Chart {idx + 1}")
        for row in data:
            chart_sheet.append(row)

        bar_chart = BarChart()
        bar_chart.title = course
        bar_chart.x_axis.title = 'Категория'
        bar_chart.y_axis.title = 'Количество студентов'
        categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=5)
        values = Reference(chart_sheet, min_col=2, min_row=2, max_row=5)
        bar_chart.add_data(values)
        bar_chart.set_categories(categories)
        bar_chart.shape = 4

        sheet.add_chart(bar_chart, f'A{row_num + idx * 15 + 2}')

    report_file = f"assessment_report_{assessment_number}.xlsx"
    workbook.save(report_file)  # Сохранение отчета в файл
    return report_file  # Возврат имени файла отчета


# Функция для получения данных пользователя (адвайзера) по email и паролю
def fetch_user(email, password):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(
        "SELECT * FROM advisors WHERE email = %s AND password = %s", (email, password)
    )
    user = cursor.fetchone()    # Получение результата запроса
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения
    return user # Возврат данных пользователя


# Функция для проверки наличия email в базе данных
def check_email(email):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(
        "SELECT * FROM advisors WHERE email = %s", (email,)
    )   # Запрос на проверку email
    user = cursor.fetchone()    # Получение результата запроса
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения
    return user # Возврат данных пользователя (если существует)


# Функция для получения списка студентов для конкретного адвайзера
def fetch_students(advisor_id):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(
        "SELECT * FROM students WHERE advisor_id = %s", (advisor_id,)
    )   # Запрос списка студентов для конкретного адвайзера
    students = cursor.fetchall()    # Получение всех результатов
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения
    return students # Возврат списка студентов


# Функция для генерации Excel-отчета для адвайзера
def generate_excel_report(advisor_id):
    students = fetch_students(advisor_id)       # Получение списка студентов для конкретного адвайзера
    workbook = openpyxl.Workbook()      # Создание новой рабочей книги
    sheet = workbook.active     # Получение активного листа
    sheet.title = 'Student Report'      # Установка названия листа

    # Заголовки
    headers = ['№', 'ФИО', 'Курс', 'ИИН', 'Email', 'Телефон']
    sheet.append(headers)   # Добавление заголовков на лист

    # Стиль заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Заполнение данных студентов
    row_num = 2
    for idx, student in enumerate(students, start=1):
        sheet.append(
            [idx, f"{student['surname']} {student['name']}", 4, student['iin'], student['email'], student['phone']])

        # Стиль данных
        for col_num in range(1, 7):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Установка цвета фона для строк
        if row_num % 2 == 0:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for col_num in range(1, 7):
            sheet.cell(row=row_num, column=col_num).fill = fill

        row_num += 1

    # Авторазмер колонок
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    report_file = f"report_{advisor_id}.xlsx"
    workbook.save(report_file)  # Сохранение отчета в файл
    return report_file  # Возврат имени файла отчета


# Класс для генерации PDF-отчета
class PDF(FPDF):
    pass


# Функция для генерации PDF-отчета по ИУП студента
def generate_pdf_isp(student_name):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute(
        """
        SELECT s.name, s.surname, c.code, c.name as course_name, c.credits, c.semester
        FROM students s
        JOIN student_semester_courses ssc ON s.student_id = ssc.student_id
        JOIN courses c ON ssc.course_id = c.course_id
        WHERE s.name || ' ' || s.surname = %s
        ORDER BY c.semester
        """, (student_name,)
    )
    results = cursor.fetchall()   # Получение всех результатов
    cursor.close()    # Закрытие курсора
    connection.close()  # Закрытие соединения

    pdf = PDF() # Создание нового PDF-документа
    pdf.add_page()  # Добавление страницы
    pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
    pdf.add_font('DejaVuB', '', 'DejaVuSansCondensed-Bold.ttf', uni=True)
    pdf.set_font('DejaVu', '', 12)

    if results:
        full_name = f"{results[0]['name']} {results[0]['surname']}"
        pdf.cell(200, 10, txt=full_name, ln=True, align='C')
        semesters = {}

        for result in results:
            semester = result['semester']
            course_details = {
                'code': result['code'],
                'name': result['course_name'],
                'credits': result['credits']
            }
            if semester not in semesters:
                semesters[semester] = []
            semesters[semester].append(course_details)

        for semester, courses in semesters.items():
            pdf.set_font("DejaVuB", '', 12)
            pdf.cell(200, 10, txt=f"Семестр {semester}", ln=True)
            pdf.set_font("DejaVu", '', 12)
            pdf.cell(60, 10, txt="Код дисциплины", border=1, ln=False)
            pdf.cell(100, 10, txt="Дисциплины", border=1, ln=False)
            pdf.cell(30, 10, txt="Кредиты", border=1, ln=True)
            for course in courses:
                pdf.cell(60, 10, txt=course['code'], border=1, ln=False)
                pdf.cell(100, 10, txt=course['name'], border=1, ln=False)
                pdf.cell(30, 10, txt=str(course['credits']), border=1, ln=True)
    else:
        pdf.cell(200, 10, txt="Студент не найден или не зарегистрирован на курсы", ln=True)

    report_file = f"ISP_{student_name}.pdf"
    pdf.output(report_file, 'F')    # Сохранение PDF-отчета в файл
    return report_file  # Возврат имени файла отчета


# Функция для получения списка студентов с проблемными оценками и пропусками
def fetch_problem_students(advisor_id):
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей
    cursor.execute("""
        SELECT s.student_id, s.surname, s.name, g.course_id, g.assessment1, g.assessment2, g.attendance, c.name as course_name
        FROM students s
        JOIN grades g ON s.student_id = g.student_id
        JOIN courses c ON g.course_id = c.course_id
        WHERE (g.assessment1 <= 15 OR g.assessment2 <= 15 OR g.attendance >= 15) AND s.advisor_id = %s
    """, (advisor_id,))     # Запрос студентов с проблемными оценками и пропусками
    problem_students = cursor.fetchall()    # Получение всех результатов
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения

    students = {}
    for entry in problem_students:
        student_id = entry['student_id']
        if student_id not in students:
            students[student_id] = {
                'surname': entry['surname'],
                'name': entry['name'],
                'assessments': [],
                'attendances': []
            }
        if entry['assessment1'] <= 15:
            students[student_id]['assessments'].append({
                'course_name': entry['course_name'],
                'score': entry['assessment1'],
                'assessment_number': 1
            })
        if entry['assessment2'] <= 15:
            students[student_id]['assessments'].append({
                'course_name': entry['course_name'],
                'score': entry['assessment2'],
                'assessment_number': 2
            })
        if entry['attendance'] >= 15:
            students[student_id]['attendances'].append({
                'course_name': entry['course_name'],
                'percentage': entry['attendance']
            })

    return list(students.values())  # Возврат списка студентов с проблемными оценками и пропусками


# Функция для получения списка студентов с проблемами в выборе элективов
def fetch_problem_students_based_on_electives():
    connection = get_db_connection()    # Получение соединения с БД
    cursor = connection.cursor(cursor_factory=RealDictCursor)   # Создание курсора с фабрикой словарей


    cursor.execute("""
        SELECT s.student_id, s.surname, s.name, c.code, c.name as course_name, eb.block_number
        FROM students s
        JOIN student_semester_courses ssc ON s.student_id = ssc.student_id
        JOIN courses c ON ssc.course_id = c.course_id
        JOIN elective_courses ec ON c.course_id = ec.course_id
        JOIN elective_blocks eb ON ec.block_id = eb.block_id
    """)    # Запрос студентов с проблемами в выборе элективов

    student_courses = cursor.fetchall()     # Получение всех результатов
    cursor.close()  # Закрытие курсора
    connection.close()  # Закрытие соединения

    # Идентифицируем студентов, которые выбрали более одного курса из одного и того же элективного блока
    student_electives = {}
    for entry in student_courses:
        student_id = entry['student_id']
        block_number = entry['block_number']
        course_info = f"{entry['code']} – {entry['course_name']}"

        if student_id not in student_electives:
            student_electives[student_id] = {
                'surname': entry['surname'],
                'name': entry['name'],
                'blocks': {}
            }
        if block_number not in student_electives[student_id]['blocks']:
            student_electives[student_id]['blocks'][block_number] = []

        student_electives[student_id]['blocks'][block_number].append(course_info)

    problem_students = []
    for student_id, student_info in student_electives.items():
        for block_number, courses in student_info['blocks'].items():
            if len(courses) > 1:
                problem_students.append({
                    'surname': student_info['surname'],
                    'name': student_info['name'],
                    'courses': courses,
                    'block_number': block_number
                })

    return problem_students     # Возврат списка студентов с проблемами в выборе элективов


